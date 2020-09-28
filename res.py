import openpyxl

import math


wb = openpyxl.load_workbook('dose.xlsx')

# print(type(wb))

# print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet1')

# print(sheet)

# print(type(sheet))

# print(sheet.title)

# print(sheet['A1'])

# print(sheet['A1'].value)
# NAME

c = sheet['B1']

# print('Row '+str(c.row)+', Column '+str(c.column)+' is '+str(c.value))

# Row 1, Column 2 is DOSE

# print('Cell ' + c.coordinate + ' is ' + c.value)

#Cell B1 is DOSE



def display(dist, res, zone1, extent, speed):

    radiation = (37.5*zone1*zone1)/(dist*dist)

    walkDist = dist*2*math.tan(math.radians(extent/2))

    timeTaken = walkDist/speed

    radiationAbsorbed = radiation*timeTaken

    a = ''
    p = 0


    for i in range(2, 42):

        if int(sheet.cell(row=i, column =2).value)+radiationAbsorbed < res :

            p = p+1

            text = sheet.cell(row=i, column =1).value
            text1 = text.ljust(30)
            text2 = str(sheet.cell(row=i, column =2).value).ljust(3) + '\t'
            text3 = str(sheet.cell(row=i, column =2).value+int(radiationAbsorbed)) + '\n'

            a = a+ str(p)+'\t'+ text1+text2+text3


    return a, p
